"""
build_weekly.py — Dashboard semanal Amazon multi-mercado (v2)
=============================================================
Genera un HTML interactivo con:
  - Pestañas por mercado (ES, FR, IT, DE, NL, BE, PL, UK, SE)
  - Semáforo por semana (>=90% media=verde, 60-89%=ambar, <60%=rojo)
  - YOY con delta absoluto (+15% / +23 uds)
  - Peso % de categoria sobre total mercado
  - Tendencia: W4 vs media W1-W3
  - Tabla 3 niveles colapsable: Familia -> Subfamilia -> ASIN
  - KPIs + graficos por mercado
  - Historico acumulado embebido (grafico de tendencia semanal)
  - Catalogo ES compartido entre todos los mercados

FUENTE: MerchantSpring (generateOrderedRevenueReport, Vendor manufacturing)
CATALOGO: catalog_es.json; si no existe, lo lee desde Excel
"""

import sys, csv, os, json
sys.stdout.reconfigure(encoding='utf-8')

# ===========================================================
#  CONFIG
# ===========================================================

CLIENT_NAME   = "Finocam"
CLIENT_LETTER = "F"

MARKET_CONFIGS = [
    {'code': 'ES', 'flag': '🇪🇸', 'label': 'España',        'dir': '',   'currency': '€',  'tz': 'Europe/Madrid'},
    {'code': 'FR', 'flag': '🇫🇷', 'label': 'France',         'dir': 'fr', 'currency': '€',  'tz': 'Europe/Paris'},
    {'code': 'IT', 'flag': '🇮🇹', 'label': 'Italia',         'dir': 'it', 'currency': '€',  'tz': 'Europe/Rome'},
    {'code': 'DE', 'flag': '🇩🇪', 'label': 'Deutschland',    'dir': 'de', 'currency': '€',  'tz': 'Europe/Berlin'},
    {'code': 'NL', 'flag': '🇳🇱', 'label': 'Nederland',      'dir': 'nl', 'currency': '€',  'tz': 'Europe/Amsterdam'},
    {'code': 'BE', 'flag': '🇧🇪', 'label': 'Belgique',       'dir': 'be', 'currency': '€',  'tz': 'Europe/Brussels'},
    {'code': 'PL', 'flag': '🇵🇱', 'label': 'Polska',         'dir': 'pl', 'currency': 'zł', 'tz': 'Europe/Warsaw'},
    {'code': 'UK', 'flag': '🇬🇧', 'label': 'United Kingdom', 'dir': 'uk', 'currency': '£',  'tz': 'Europe/London'},
    {'code': 'SE', 'flag': '🇸🇪', 'label': 'Sverige',        'dir': 'se', 'currency': 'kr', 'tz': 'Europe/Stockholm'},
]

_SCRIPT_DIR  = os.path.dirname(os.path.abspath(__file__))
CSV_BASE     = os.path.join(_SCRIPT_DIR, 'data')
CATALOG_JSON = os.path.join(_SCRIPT_DIR, 'data', 'catalog_es.json')
HISTORY_JSON = os.path.join(_SCRIPT_DIR, 'data', 'history.json')
EXCEL_PATH   = os.environ.get('WEEKLY_EXCEL', r'C:\Users\Daniela\Downloads\FINOCAM_Familias_Subfamilias_ASIN.xlsx')
OUT_PATH     = os.environ.get('WEEKLY_OUT_PATH', os.path.join(os.path.dirname(_SCRIPT_DIR), 'FINOCAM_Weekly.html'))

# Hojas "coleccion" (estructura diferente en Excel)
COLS = {'Moniquilla', 'Talkual'}
SKIP = {'Resumen'}

def is_valid_anualidad(val):
    s = str(val or '').strip()
    return s.startswith('2026') or s.startswith('2027') or s == 'NOCAD'

CAT_ORDER = [
    'Moniquilla', 'Talkual',
    'Agendas', 'Calendarios', 'Cuadernos',
    'Planificadores', 'Libros de Firma', 'Indices',
    'Portadocumentos', 'Recambios',
    '_RESTO_',
]

# Semanas
import datetime as _dt
_MONTHS_ES = ['','Ene','Feb','Mar','Abr','May','Jun','Jul','Ago','Sep','Oct','Nov','Dic']

def _auto_weeks():
    today = _dt.date.today()
    this_mon = today - _dt.timedelta(days=today.weekday())
    out = []
    for i in range(4, 0, -1):
        mon = this_mon - _dt.timedelta(weeks=i)
        sun = mon + _dt.timedelta(days=6)
        if mon.month == sun.month:
            label = f'{mon.day}–{sun.day} {_MONTHS_ES[mon.month]}'
        else:
            label = f'{mon.day} {_MONTHS_ES[mon.month]}–{sun.day} {_MONTHS_ES[sun.month]}'
        iso_wk = mon.isocalendar()[1]
        out.append({'label': label, 'wk': f'W{5-i}', 'iso_wk': iso_wk, 'iso_yr': mon.year})
    return out

WEEKS = _auto_weeks() if os.environ.get('WEEKLY_AUTO') == '1' else [
    {'label': '13–19 Jul', 'wk': 'W1', 'iso_wk': 29, 'iso_yr': 2026},
    {'label': '20–26 Jul', 'wk': 'W2', 'iso_wk': 30, 'iso_yr': 2026},
    {'label': '27 Jul–2 Ago', 'wk': 'W3', 'iso_wk': 31, 'iso_yr': 2026},
    {'label': '3–9 Ago',  'wk': 'W4', 'iso_wk': 32, 'iso_yr': 2026},
]

UPDATE_DATE = _dt.date.today().isoformat()

# ===========================================================
#  1. CATALOGO
# ===========================================================

if os.path.exists(CATALOG_JSON):
    with open(CATALOG_JSON, encoding='utf-8') as f:
        catalog = json.load(f)
    print(f'Catalogo: {len(catalog)} ASINs (catalog_es.json)')
else:
    import openpyxl
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    catalog = {}
    for sname in wb.sheetnames:
        if sname in SKIP or sname in COLS:
            continue
        for r in list(wb[sname].iter_rows(values_only=True))[1:]:
            if not r[1]:
                continue
            asin = str(r[1]).strip()
            if not is_valid_anualidad(r[2]):
                continue
            if asin not in catalog:
                catalog[asin] = {'familia': sname, 'sub': str(r[0] or '').strip(),
                                 'any': str(r[2] or '').strip(), 'desc': str(r[3] or '').strip(), 'col': None}
    for col in COLS:
        if col not in wb.sheetnames:
            continue
        for r in list(wb[col].iter_rows(values_only=True))[1:]:
            if not r[2]:
                continue
            asin = str(r[2]).strip()
            if not is_valid_anualidad(r[3]):
                continue
            entry = {'familia': str(r[0] or '').strip(), 'sub': str(r[1] or '').strip(),
                     'any': str(r[3] or '').strip(), 'desc': str(r[4] or '').strip(), 'col': col}
            if asin not in catalog:
                catalog[asin] = entry
            else:
                catalog[asin]['col'] = col
    print(f'Catalogo: {len(catalog)} ASINs (Excel)')

# ===========================================================
#  2. CARGA DE CSVs POR MERCADO
# ===========================================================

def load_market_csvs(market_dir):
    """Carga los 4 CSVs semanales de un mercado. Devuelve aw, titles."""
    aw = {}
    titles = {}
    dir_path = os.path.join(CSV_BASE, market_dir) if market_dir else CSV_BASE
    for wi in range(4):
        csv_path = os.path.join(dir_path, f'weekly_{wi}.csv')
        if not os.path.exists(csv_path):
            continue
        with open(csv_path, encoding='utf-8-sig') as f:
            for row in csv.DictReader(f):
                asin = row['asin']
                if asin not in catalog:
                    continue
                try:
                    rev = float(row['orderedRevenue'])
                    u   = int(row['orderedUnits'])
                    pr  = float(row['priorOrderedRevenue'])
                    pu  = int(row['priorOrderedUnits'])
                    avg_p = float(row.get('avgPrice') or 0)
                except (ValueError, KeyError):
                    continue
                aw.setdefault(asin, {})[wi] = {'rev': rev, 'u': u, 'pr': pr, 'pu': pu, 'avg': avg_p}
                titles[asin] = row.get('title', '')
    return aw, titles

# ===========================================================
#  3. AGREGACION
# ===========================================================

def agg(asins, aw):
    W = [{'rev': 0, 'u': 0, 'pr': 0, 'pu': 0, 'has': False} for _ in range(4)]
    for a in asins:
        for wi in range(4):
            d = (aw.get(a) or {}).get(wi)
            if d and (d['u'] > 0 or d['rev'] > 0):
                W[wi]['rev'] += d['rev']
                W[wi]['u']   += d['u']
                W[wi]['pr']  += d['pr']
                W[wi]['pu']  += d['pu']
                W[wi]['has']  = True
    total_u   = sum(w['u']   for w in W)
    total_r   = sum(w['rev'] for w in W)
    days      = sum(7 for w in W if w['has'])
    mu = total_u / days if days else 0
    mr = total_r / days if days else 0
    prev3 = [W[i] for i in range(3) if W[i]['has']]
    t = None
    if W[3]['has'] and prev3:
        avg = sum(w['u'] for w in prev3) / len(prev3)
        if avg > 0:
            t = (W[3]['u'] - avg) / avg * 100
    return {'W': W, 'mu': mu, 'mr': mr, 'tend': t, 'total_u': total_u, 'total_r': total_r}

# ===========================================================
#  4. ESTRUCTURA POR MERCADO
# ===========================================================

def build_struct(catalog):
    struct = {}
    for asin, m in catalog.items():
        cat = m['col'] or m['familia']
        struct.setdefault(cat, {}).setdefault(m['sub'], [])
        if asin not in struct[cat][m['sub']]:
            struct[cat][m['sub']].append(asin)
    EXPLICIT_CATS = set(CAT_ORDER) - {'_RESTO_'}
    resto_subs = {}
    for cat_key, subs in struct.items():
        if cat_key not in EXPLICIT_CATS:
            for sub, asins in subs.items():
                resto_subs.setdefault(f'{cat_key} — {sub}', []).extend(asins)
    struct['_RESTO_'] = resto_subs
    return struct

full_struct = build_struct(catalog)

def build_dash(aw, currency):
    dash = {}
    for cat in CAT_ORDER:
        if cat not in full_struct:
            continue
        display_name = 'Resto' if cat == '_RESTO_' else cat
        all_a = [a for subs in full_struct[cat].values() for a in subs]
        if not any(a in aw for a in all_a):
            continue
        ca = agg(all_a, aw)
        subs_d = {}
        for sub, asins in sorted(full_struct[cat].items()):
            if not any(a in aw for a in asins):
                continue
            sa = agg(asins, aw)
            ad = {}
            for a in sorted(asins, key=lambda x: -(aw.get(x, {}).get(3, {'u': 0})['u'])):
                if a not in aw:
                    continue
                # avg price across weeks
                total_u_a = sum((aw[a].get(wi) or {}).get('u', 0) for wi in range(4))
                total_r_a = sum((aw[a].get(wi) or {}).get('rev', 0.0) for wi in range(4))
                avg_price = total_r_a / total_u_a if total_u_a > 0 else 0
                ad[a] = {
                    'desc':      (catalog[a]['desc'] or '')[:48],
                    'agg':       agg([a], aw),
                    'avg_price': avg_price,
                }
            if ad:
                subs_d[sub] = {'agg': sa, 'asins': ad}
        if subs_d:
            dash[cat] = {'agg': ca, 'subs': subs_d, 'display': display_name}
    return dash

# Procesar cada mercado
all_market_data = {}
for mc in MARKET_CONFIGS:
    code = mc['code']
    aw, titles = load_market_csvs(mc['dir'])
    n_active = sum(1 for a in aw if any(aw[a].get(wi, {}).get('u', 0) > 0 for wi in range(4)))
    print(f'  {mc["flag"]} {code}: {n_active} ASINs con ventas')
    if n_active == 0:
        all_market_data[code] = None
        continue
    dash = build_dash(aw, mc['currency'])
    week_totals_u = [sum(dash[cat]['agg']['W'][wi]['u'] for cat in dash) for wi in range(4)]
    week_totals_r = [sum(dash[cat]['agg']['W'][wi]['rev'] for cat in dash) for wi in range(4)]
    cat_chart = []
    for cat in CAT_ORDER:
        if cat not in dash:
            continue
        d = dash[cat]['agg']
        wu = [d['W'][i]['u'] for i in range(4)]
        wr = [round(d['W'][i]['rev'], 0) for i in range(4)]
        n_asins = sum(len(subs['asins']) for subs in dash[cat]['subs'].values())
        cat_chart.append({'name': dash[cat].get('display', cat), 'u': wu, 'r': wr, 'active': n_asins})
    total_u = sum(week_totals_u)
    total_r = sum(week_totals_r)
    all_market_data[code] = {
        'aw': aw, 'titles': titles, 'dash': dash,
        'week_totals_u': week_totals_u, 'week_totals_r': week_totals_r,
        'total_u': total_u, 'total_r': total_r,
        'cat_chart': cat_chart,
        'currency': mc['currency'],
        'n_active': n_active,
        'mc': mc,
    }

# ===========================================================
#  5. HISTORICO
# ===========================================================

history = {}
if os.path.exists(HISTORY_JSON):
    try:
        with open(HISTORY_JSON, encoding='utf-8') as f:
            history = json.load(f)
    except Exception:
        history = {}

# Guardar datos de cada semana del run actual
for wi, wk in enumerate(WEEKS):
    key = f'{wk["iso_yr"]}_W{wk["iso_wk"]:02d}'
    entry = {'label': wk['label'], 'iso': wk['iso_wk'], 'yr': wk['iso_yr'], 'markets': {}}
    for mc in MARKET_CONFIGS:
        md = all_market_data.get(mc['code'])
        if md is None:
            entry['markets'][mc['code']] = None
            continue
        cats = {}
        for cat in CAT_ORDER:
            if cat not in md['dash']:
                continue
            d = md['dash'][cat]['agg']
            cats[md['dash'][cat].get('display', cat)] = {
                'u': d['W'][wi]['u'], 'r': round(d['W'][wi]['rev'], 1)
            }
        entry['markets'][mc['code']] = {
            'u': md['week_totals_u'][wi],
            'r': round(md['week_totals_r'][wi], 1),
            'cats': cats,
        }
    history[key] = entry

# Ordenar por clave (cronologico)
history = dict(sorted(history.items()))

with open(HISTORY_JSON, 'w', encoding='utf-8') as f:
    json.dump(history, f, ensure_ascii=False)
print(f'Historico: {len(history)} semanas guardadas')

# ===========================================================
#  6. HELPERS HTML
# ===========================================================

def fu(v):
    if v == 0: return '—'
    if v >= 1000: return f'{v/1000:.1f}k'
    return str(int(round(v)))

def fr_val(v, currency='€'):
    if v == 0: return ''
    sym = currency
    if v >= 1000: return f'{v/1000:.1f}k{sym}'
    return f'{int(round(v))}{sym}'

def fmed(u): return f'{int(round(u))}/d' if u > 0 else '—'

def tend_html(t):
    if t is None:   return '<span class="t-na">—</span>'
    if abs(t) < 3:  return f'<span class="t-flat">≈{t:+.0f}%</span>'
    if t > 0:       return f'<span class="t-up">↑{t:.0f}%</span>'
    return f'<span class="t-dn">↓{abs(t):.0f}%</span>'

def yoy_span(cur, prev, cur_label, prev_label, cls_prefix):
    if not prev or prev < 5 or not cur: return ''
    pct = (cur - prev) / prev * 100
    if abs(pct) > 499: return ''
    delta = int(round(cur - prev))
    sign = '+' if pct >= 0 else ''
    dsign = '+' if delta >= 0 else ''
    cls = 'yb-up' if pct >= 0 else 'yb-dn'
    return (f'<span class="{cls_prefix} {cls}" title="vs {prev_label} año anterior">'
            f'{sign}{pct:.0f}%&nbsp;<small>({dsign}{delta})</small></span>')

def weight_badge(cat_u, total_u):
    if not total_u: return ''
    pct = cat_u / total_u * 100
    return f'<span class="wpct">{pct:.1f}%</span>'

def week_cells(a, total_u, currency='€'):
    W = a['W']; mu = a['mu']
    out = ''
    for w in W:
        cc = 'cn'
        if w['u'] > 0 and mu > 0:
            p  = w['u'] / mu
            cc = 'cg' if p >= 0.9 else ('cy' if p >= 0.6 else 'cr')
        yoy_u = yoy_span(w['u'],   w['pu'], 'uds', 'uds', 'yy-u')
        yoy_r = yoy_span(w['rev'], w['pr'], currency, currency, 'yy-r')
        yoy_block = f'<div class="wy">{yoy_u}{yoy_r}</div>' if (yoy_u or yoy_r) else ''
        out += f'''<td class="wk"><div class="wi {cc}">
          <div class="wu">{fu(w["u"])}</div>
          <div class="wr">{fr_val(w["rev"], currency)}</div>
          {yoy_block}
        </div></td>'''
    out += f'<td class="tend">{tend_html(a["tend"])}</td>'
    return out

def build_rows(dash, aw, total_u, currency='€'):
    html = ''
    for cat in CAT_ORDER:
        if cat not in dash: continue
        d = dash[cat]; is_col = cat in COLS
        display = d.get('display', cat)
        n_a = sum(len(v['asins']) for v in d['subs'].values())
        cat_id = cat.replace(' ', '_').replace('/', '_')
        wb = weight_badge(d['agg']['total_u'], total_u)
        html += f'''<tr class="rc {'col' if is_col else ''}" data-cat="{cat_id}">
          <td class="nc" onclick="tgC(this)">
            <span class="ci">{'◆' if is_col else '▸'}</span>
            <span class="cn-lbl">{display}</span>
            {wb}
            <span class="meta">{len(d["subs"])} subfam · {n_a} ASINs</span>
          </td>
          <td class="mc"><div class="med-pill">{fmed(d["agg"]["mu"])}</div></td>
          {week_cells(d["agg"], total_u, currency)}
        </tr>'''
        for sub, sd in d['subs'].items():
            sub_id = sub.replace("'", "").replace('"', '').replace(' ', '_')
            html += f'''<tr class="rs" data-p="{cat_id}" style="display:none">
              <td class="ns" onclick="tgS(this)">
                <span class="si">▹</span><span class="sn-lbl">{sub}</span>
                <span class="meta">{len(sd["asins"])} ASINs</span>
              </td>
              <td class="mc"><div class="med-pill sm">{fmed(sd["agg"]["mu"])}</div></td>
              {week_cells(sd["agg"], total_u, currency)}
            </tr>'''
            sub_total_u = sd['agg']['total_u'] or 1
            for asin, ad in sd['asins'].items():
                asin_u = ad['agg']['total_u']
                bar_pct = min(100, round(asin_u / sub_total_u * 100)) if sub_total_u else 0
                price_str = f'{ad["avg_price"]:.2f}{currency}' if ad['avg_price'] > 0 else ''
                html += f'''<tr class="ra" data-p="{cat_id}" data-s="{sub_id}" style="display:none">
                  <td class="an" title="{asin} — {ad["desc"]}">
                    <span class="ac">{asin}</span>
                    <span class="adesc">{ad["desc"]}</span>
                    <div class="abar-wrap"><div class="abar" style="width:{bar_pct}%"></div></div>
                    {f'<span class="aprice">{price_str}</span>' if price_str else ''}
                  </td>
                  <td class="mc"><span class="med-u">{fmed(ad["agg"]["mu"])}</span></td>
                  {week_cells(ad["agg"], total_u, currency)}
                </tr>'''
    return html

# ===========================================================
#  7. HTML POR MERCADO
# ===========================================================

def market_section(code, md, WEEKS):
    if md is None:
        return f'<div id="mkt-{code}" class="mkt-section" style="display:none"><div class="no-data">Sin datos disponibles para {code} en este periodo</div></div>'

    mc       = md['mc']
    currency = mc['currency']
    total_u  = md['total_u']
    total_r  = md['total_r']
    wtU      = md['week_totals_u']
    wtR      = md['week_totals_r']
    best_wi  = wtU.index(max(wtU)) if any(v > 0 for v in wtU) else 0
    WEEK_NAMES = [w['label'] for w in WEEKS]

    chart_data = json.dumps({
        'weekNames': WEEK_NAMES,
        'weekTotalsU': wtU,
        'weekTotalsR': [round(v) for v in wtR],
        'cats': md['cat_chart'],
    })

    wk_ths = ''.join(
        f'<th class="wkh"><div class="wkl">{w["wk"]}<sup class="iso-w">W{w["iso_wk"]}</sup></div><div class="wkd">{w["label"]}</div></th>'
        for w in WEEKS
    )
    rows_html = build_rows(md['dash'], md['aw'], total_u, currency)

    total_r_fmt = fr_val(total_r, currency)
    avg_day_r   = fr_val(total_r / 28, currency)
    best_wk_lbl = WEEK_NAMES[best_wi]
    best_wk_u   = round(wtU[best_wi])

    return f'''<div id="mkt-{code}" class="mkt-section" style="display:none">
  <div class="summary">
    <div class="kpi-row">
      <div class="kpi kpi-acc">
        <div class="kpi-val">{round(total_u):,}</div>
        <div class="kpi-lbl">Uds vendidas · 4 semanas</div>
        <div class="kpi-sub">Mejor: {best_wk_lbl} ({best_wk_u:,} uds)</div>
      </div>
      <div class="kpi">
        <div class="kpi-val">{total_r_fmt}</div>
        <div class="kpi-lbl">Facturación · 4 semanas</div>
        <div class="kpi-sub">Media/día: {avg_day_r}</div>
      </div>
      <div class="kpi kpi-g">
        <div class="kpi-val">{md["n_active"]:,}</div>
        <div class="kpi-lbl">ASINs activos</div>
        <div class="kpi-sub">{len(catalog)-md["n_active"]:,} sin ventas</div>
      </div>
    </div>
    <div class="chart-area">
      <div class="chart-box">
        <div class="chart-title">Evolución semanal · unidades</div>
        <canvas id="cvWeek-{code}" height="110"></canvas>
      </div>
      <div class="chart-box">
        <div class="chart-title">Tendencia por categoría · unidades</div>
        <canvas id="cvCat-{code}" height="110"></canvas>
      </div>
    </div>
  </div>
  <div class="toolbar">
    <span class="tlbl">Semáforo</span>
    <div class="leg"><div class="leg-b" style="background:var(--g)"></div>≥90%</div>
    <div class="leg"><div class="leg-b" style="background:var(--y)"></div>60–89%</div>
    <div class="leg"><div class="leg-b" style="background:var(--r)"></div>&lt;60%</div>
    <div class="tbar-sep"></div>
    <button class="btn-sm on" id="bRev-{code}" onclick="togRev('{code}')">€ Facturación</button>
    <button class="btn-sm on" id="bU-{code}"   onclick="togU('{code}')">📦 Unidades</button>
    <div class="tbar-sep"></div>
    <button class="btn-sm" onclick="expAll('{code}')">+ Todo</button>
    <button class="btn-sm" onclick="colAll('{code}')">− Todo</button>
    <span style="margin-left:auto;font-size:9px;color:var(--t3)">W=ISO week · Tendencia: W4 vs media W1–W3</span>
  </div>
  <div class="wrap">
  <table>
    <thead><tr>
      <th class="nh">Categoría / Subfamilia / ASIN</th>
      <th class="mh">Med/día</th>
      {wk_ths}
      <th class="tndh">Tend.</th>
    </tr></thead>
    <tbody id="tb-{code}">{rows_html}</tbody>
  </table>
  </div>
  <script>
  (function(){{
    var CD_{code} = {chart_data};
    var sRev_{code}=true, sU_{code}=true;
    function togRev_{code}(){{
      sRev_{code}=!sRev_{code};
      var b=document.getElementById('bRev-{code}');
      if(b)b.classList.toggle('on',sRev_{code});
      var sec=document.getElementById('mkt-{code}');
      sec.querySelectorAll('.wr').forEach(function(e){{e.style.display=sRev_{code}?'':'none';}});
      sec.querySelectorAll('.yy-r').forEach(function(e){{e.style.display=(!sU_{code}&&sRev_{code})?'':'none';}});
    }}
    function togU_{code}(){{
      sU_{code}=!sU_{code};
      var b=document.getElementById('bU-{code}');
      if(b)b.classList.toggle('on',sU_{code});
      var sec=document.getElementById('mkt-{code}');
      sec.querySelectorAll('.wu').forEach(function(e){{e.style.display=sU_{code}?'':'none';}});
      sec.querySelectorAll('.yy-u').forEach(function(e){{e.style.display=sU_{code}?'':'none';}});
      sec.querySelectorAll('.yy-r').forEach(function(e){{e.style.display=(!sU_{code}&&sRev_{code})?'':'none';}});
    }}
    window._mktTogRev=window._mktTogRev||{{}};
    window._mktTogU=window._mktTogU||{{}};
    window._mktTogRev['{code}']=togRev_{code};
    window._mktTogU['{code}']=togU_{code};
    window.CHART_DATA=window.CHART_DATA||{{}};
    window.CHART_DATA['{code}']=CD_{code};
  }})();
  </script>
</div>'''

# ===========================================================
#  8. HISTORICO JS DATA
# ===========================================================

hist_data = []
for key, entry in history.items():
    row = {'key': key, 'label': entry.get('label', key), 'iso': entry.get('iso', 0), 'yr': entry.get('yr', 2026), 'markets': {}}
    for mc in MARKET_CONFIGS:
        code = mc['code']
        mkt = (entry.get('markets') or {}).get(code)
        row['markets'][code] = mkt['u'] if mkt else 0
    hist_data.append(row)

HIST_JS = json.dumps(hist_data)
MKT_JS  = json.dumps([{'code': mc['code'], 'flag': mc['flag'], 'label': mc['label']} for mc in MARKET_CONFIGS])

# ===========================================================
#  9. HTML COMPLETO
# ===========================================================

market_tabs_html = ''.join(
    f'<button class="btn-mkt" id="tab-{mc["code"]}" onclick="showMkt(\'{mc["code"]}\')">'
    f'{mc["flag"]} {mc["code"]}</button>'
    for mc in MARKET_CONFIGS
)

all_sections_html = '\n'.join(market_section(mc['code'], all_market_data.get(mc['code']), WEEKS) for mc in MARKET_CONFIGS)

# find first market with data for default tab
default_mkt = next((mc['code'] for mc in MARKET_CONFIGS if all_market_data.get(mc['code']) is not None), 'ES')

html = f'''<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Finocam · Seguimiento Semanal Amazon</title>
<style>
:root{{
  --bg:#090c14;--s1:#0f1320;--s2:#151929;--s3:#1d2235;--s4:#252b40;
  --bd:#2a2f4a;--bd2:#1a1f33;
  --t1:#dde3f5;--t2:#8892b0;--t3:#4a5270;
  --acc:#ff9900;--acc2:#ffb340;
  --g:#22c55e;--gb:rgba(34,197,94,.09);--gbd:rgba(34,197,94,.4);
  --y:#f59e0b;--yb:rgba(245,158,11,.09);--ybd:rgba(245,158,11,.4);
  --r:#ef4444;--rb:rgba(239,68,68,.09);--rbd:rgba(239,68,68,.4);
  --ff:'Segoe UI',system-ui,-apple-system,sans-serif;
  --ff-mono:'Cascadia Code','Consolas',monospace;
}}
@media(prefers-color-scheme:light){{:root{{
  --bg:#f0f2f8;--s1:#fff;--s2:#f7f9fe;--s3:#edf0f8;--s4:#e4e8f4;
  --bd:#cdd2e8;--bd2:#dde1f0;--t1:#131728;--t2:#556080;--t3:#8898b8;
}}}}
:root[data-theme="dark"]{{--bg:#090c14;--s1:#0f1320;--s2:#151929;--s3:#1d2235;--s4:#252b40;--bd:#2a2f4a;--bd2:#1a1f33;--t1:#dde3f5;--t2:#8892b0;--t3:#4a5270;}}
:root[data-theme="light"]{{--bg:#f0f2f8;--s1:#fff;--s2:#f7f9fe;--s3:#edf0f8;--s4:#e4e8f4;--bd:#cdd2e8;--bd2:#dde1f0;--t1:#131728;--t2:#556080;--t3:#8898b8;}}
*{{box-sizing:border-box;margin:0;padding:0}}
body{{background:var(--bg);color:var(--t1);font-family:var(--ff);font-size:12px}}
/* HEADER */
.hdr{{background:var(--s2);border-bottom:1px solid var(--bd);padding:10px 20px;display:flex;align-items:center;gap:14px;flex-wrap:wrap}}
.hdr-logo{{width:32px;height:32px;border-radius:7px;background:linear-gradient(135deg,#ff9900,#e07000);display:flex;align-items:center;justify-content:center;font-size:16px;font-weight:900;color:#fff;flex-shrink:0;box-shadow:0 2px 10px rgba(255,153,0,.25)}}
.hdr-t1{{font-size:14px;font-weight:700;color:var(--t1)}}
.hdr-t2{{font-size:10px;color:var(--t3);margin-top:1px}}
.hdr-right{{margin-left:auto;display:flex;align-items:center;gap:8px}}
.upd-badge{{font-size:9px;color:var(--t3)}}
.btn-sm{{background:var(--s3);border:1px solid var(--bd);color:var(--t2);padding:4px 9px;border-radius:5px;cursor:pointer;font-size:11px;transition:all .15s}}
.btn-sm:hover{{background:var(--s4);color:var(--t1)}}
.btn-sm.on{{background:rgba(255,153,0,.13);border-color:rgba(255,153,0,.35);color:var(--acc)}}
/* MARKET TABS */
.mkt-nav{{background:var(--s1);border-bottom:2px solid var(--bd);padding:0 20px;display:flex;gap:2px;overflow-x:auto}}
.btn-mkt{{border:none;background:transparent;color:var(--t2);font-size:12px;font-weight:600;padding:10px 14px;cursor:pointer;border-bottom:2px solid transparent;margin-bottom:-2px;transition:all .15s;white-space:nowrap}}
.btn-mkt:hover{{color:var(--t1);background:var(--s2)}}
.btn-mkt.active{{color:var(--acc);border-bottom-color:var(--acc);background:rgba(255,153,0,.07)}}
/* SUMMARY */
.mkt-section{{}}
.summary{{background:var(--s1);border-bottom:1px solid var(--bd);padding:16px 20px;display:grid;gap:14px}}
.kpi-row{{display:flex;gap:10px;flex-wrap:wrap}}
.kpi{{background:var(--s2);border:1px solid var(--bd);border-radius:8px;padding:10px 14px;min-width:130px;flex:1}}
.kpi-val{{font-size:22px;font-weight:700;color:var(--t1);font-variant-numeric:tabular-nums;letter-spacing:-.5px;line-height:1}}
.kpi-lbl{{font-size:10px;color:var(--t3);text-transform:uppercase;letter-spacing:.6px;margin-top:4px}}
.kpi-sub{{font-size:10px;color:var(--t2);margin-top:2px}}
.kpi-acc .kpi-val{{color:var(--acc2)}}
.kpi-g .kpi-val{{color:var(--g)}}
.chart-area{{display:grid;grid-template-columns:1fr 1fr;gap:14px}}
.chart-box{{background:var(--s2);border:1px solid var(--bd);border-radius:8px;padding:12px}}
.chart-title{{font-size:10px;font-weight:600;color:var(--t3);text-transform:uppercase;letter-spacing:.6px;margin-bottom:10px}}
canvas{{display:block;width:100%}}
/* HISTORICO */
.hist-section{{background:var(--s1);border-bottom:1px solid var(--bd);padding:16px 20px}}
.hist-header{{display:flex;align-items:center;gap:10px;margin-bottom:12px}}
.hist-title{{font-size:11px;font-weight:700;color:var(--t2);text-transform:uppercase;letter-spacing:.6px}}
.hist-mkt-sel{{display:flex;gap:4px;flex-wrap:wrap}}
.btn-hist{{background:var(--s3);border:1px solid var(--bd);color:var(--t2);padding:3px 9px;border-radius:4px;cursor:pointer;font-size:10px;transition:all .15s}}
.btn-hist.on{{background:rgba(255,153,0,.13);border-color:rgba(255,153,0,.35);color:var(--acc)}}
/* TOOLBAR */
.toolbar{{background:var(--s1);border-bottom:1px solid var(--bd2);padding:7px 20px;display:flex;align-items:center;gap:8px;flex-wrap:wrap}}
.tlbl{{font-size:9px;color:var(--t3);text-transform:uppercase;letter-spacing:.7px;font-weight:600}}
.tbar-sep{{width:1px;height:18px;background:var(--bd);margin:0 3px}}
.leg{{display:flex;align-items:center;gap:5px;font-size:10px;color:var(--t2)}}
.leg-b{{width:3px;height:12px;border-radius:1px;flex-shrink:0}}
/* TABLE */
.wrap{{overflow-x:auto}}
table{{width:100%;border-collapse:collapse;min-width:780px}}
thead th{{background:var(--s2);color:var(--t3);font-size:9px;font-weight:600;text-transform:uppercase;letter-spacing:.7px;padding:8px 10px;border-bottom:1px solid var(--bd);white-space:nowrap}}
th.nh{{text-align:left;min-width:260px}}
th.mh{{min-width:62px;text-align:center}}
th.wkh{{min-width:88px;text-align:center;border-left:1px solid var(--bd2)}}
.wkl{{font-size:11px;font-weight:700;color:var(--t2)}}
sup.iso-w{{font-size:7px;color:var(--t3);font-weight:400;margin-left:2px}}
.wkd{{font-size:9px;color:var(--t3);margin-top:1px}}
th.tndh{{min-width:68px;text-align:center}}
tr.rc td{{background:var(--s2);border-bottom:1px solid var(--bd);padding:10px 12px;transition:background .1s;cursor:pointer}}
tr.rc td.nc{{border-left:4px solid transparent}}
tr.rc:hover td{{background:var(--s3)}}
tr.rc.col td{{background:linear-gradient(90deg,rgba(255,153,0,.06) 0%,var(--s2) 50%)!important}}
tr.rc.col td.nc{{border-left:4px solid var(--acc)!important}}
tr.rc.col:hover td{{background:linear-gradient(90deg,rgba(255,153,0,.11) 0%,var(--s3) 50%)!important}}
.ci{{font-size:9px;color:var(--t3);margin-right:7px;transition:transform .15s;display:inline-block}}
.rc.open .ci{{transform:rotate(90deg);color:var(--acc)}}
.cn-lbl{{font-size:13px;font-weight:700}}
.meta{{font-size:9px;color:var(--t3);margin-left:6px}}
.nc{{display:flex;align-items:center;user-select:none;gap:0}}
.wpct{{font-size:9px;font-weight:700;background:rgba(255,153,0,.12);color:var(--acc);padding:1px 6px;border-radius:10px;margin-left:6px;flex-shrink:0}}
tr.rs td{{background:var(--s1);border-bottom:1px solid var(--bd2);padding:7px 12px 7px 26px;transition:background .1s}}
tr.rs:hover td{{background:var(--s2)}}
tr.rs td:first-child{{border-left:2px solid var(--bd)}}
.si{{font-size:9px;color:var(--t3);margin-right:5px}}
.sn-lbl{{font-size:11px;font-weight:600;color:var(--t2)}}
.ns{{display:flex;align-items:center;cursor:pointer;user-select:none}}
tr.rs.open .si{{color:var(--acc2)}}
tr.ra td{{background:var(--bg);border-bottom:1px solid rgba(26,31,51,.6);padding:5px 12px 5px 38px;vertical-align:middle}}
tr.ra:hover td{{background:var(--s1)}}
tr.ra td:first-child{{border-left:2px solid var(--bd2)}}
.an{{font-size:10px;color:var(--t2);max-width:260px}}
.ac{{font-family:var(--ff-mono);font-size:9px;color:var(--t3);margin-right:5px}}
.adesc{{display:block;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}}
.abar-wrap{{height:2px;background:var(--bd);border-radius:1px;margin-top:3px}}
.abar{{height:2px;background:var(--acc);border-radius:1px;transition:width .3s}}
.aprice{{font-size:9px;color:var(--t3);display:block;margin-top:1px}}
.mc{{text-align:center;vertical-align:middle;padding:4px 6px!important}}
.med-pill{{display:inline-block;background:var(--s3);border:1px solid var(--bd);color:var(--t2);font-size:10px;font-weight:600;padding:2px 7px;border-radius:10px;white-space:nowrap}}
.med-pill.sm{{font-size:9px;padding:1px 6px;background:var(--s2)}}
.med-u{{font-size:9px;color:var(--t3);font-weight:600}}
td.wk{{padding:2px 2px;border-left:1px solid var(--bd2);vertical-align:middle;text-align:center}}
.wi{{border-radius:6px;padding:6px 5px;display:flex;flex-direction:column;align-items:center;gap:1px;border-left:3px solid transparent;transition:background .1s}}
tr.rc td.wk .wi{{padding:8px 6px;border-radius:8px}}
.wi.cg{{background:var(--gb);border-left-color:var(--gbd)}}
.wi.cy{{background:var(--yb);border-left-color:var(--ybd)}}
.wi.cr{{background:var(--rb);border-left-color:var(--rbd)}}
.wu{{font-size:15px;font-weight:700;line-height:1;font-variant-numeric:tabular-nums}}
tr.rc .wu{{font-size:18px}}
.wi.cg .wu{{color:var(--g)}}.wi.cy .wu{{color:var(--y)}}.wi.cr .wu{{color:var(--r)}}.wi.cn .wu{{color:var(--t3)}}
.wr{{font-size:9px;color:var(--t3);font-weight:400;margin-top:1px}}
tr.rc .wr{{font-size:10px}}
.wy{{margin-top:2px;display:flex;gap:2px;justify-content:center;flex-wrap:wrap}}
.yy-u,.yy-r{{font-size:8px;font-weight:700;padding:1px 4px;border-radius:3px;white-space:nowrap}}
.yy-u small,.yy-r small{{font-size:7px;font-weight:400}}
.yy-r{{display:none}}
.yb-up{{background:rgba(34,197,94,.13);color:#4ade80}}
.yb-dn{{background:rgba(239,68,68,.11);color:#f87171}}
td.tend{{text-align:center;padding:3px 7px;vertical-align:middle}}
.t-up{{font-size:12px;font-weight:700;color:var(--g)}}
.t-dn{{font-size:12px;font-weight:700;color:var(--r)}}
.t-flat{{font-size:11px;color:var(--t3)}}
.t-na{{font-size:11px;color:var(--t3)}}
.no-data{{padding:40px 20px;text-align:center;color:var(--t3);font-size:13px}}
.ft{{padding:10px 20px;border-top:1px solid var(--bd2);font-size:9px;color:var(--t3);display:flex;gap:12px;flex-wrap:wrap}}
</style>
</head>
<body>

<div class="hdr">
  <div class="hdr-logo">F</div>
  <div>
    <div class="hdr-t1">Finocam · Seguimiento Semanal Amazon Vendor</div>
    <div class="hdr-t2">Sell-out · ASINs activos · últimas 4 semanas · 9 mercados EU · MerchantSpring</div>
  </div>
  <div class="hdr-right">
    <span class="upd-badge">Actualizado {UPDATE_DATE}</span>
    <button class="btn-sm" id="btnTh" onclick="tgTh()">🌙</button>
  </div>
</div>

<div class="mkt-nav">
  {market_tabs_html}
</div>

{all_sections_html}

<div class="hist-section" id="hist-section">
  <div class="hist-header">
    <span class="hist-title">📈 Histórico semanal</span>
    <div class="hist-mkt-sel" id="hist-mkt-btns"></div>
  </div>
  <canvas id="cvHist" height="130" style="width:100%"></canvas>
</div>

<div class="ft">
  <span>Fuente: MerchantSpring · Vendor manufacturing view</span>
  <span>·</span><span>Semáforo vs media/día del periodo</span>
  <span>·</span><span>YOY = misma semana año anterior · (Δ) = diferencia absoluta en unidades</span>
  <span>·</span><span>% = peso de categoría sobre total mercado</span>
</div>

<script>
var HIST_DATA={HIST_JS};
var MKT_LIST={MKT_JS};
var _activeMkt='{default_mkt}';
var _histMkts=['{default_mkt}'];

// Tema
function tgTh(){{
  var r=document.documentElement,c=r.getAttribute('data-theme')||'dark';
  r.setAttribute('data-theme',c==='dark'?'light':'dark');
  document.getElementById('btnTh').textContent=c==='dark'?'☀':'🌙';
  setTimeout(function(){{drawAllCharts(_activeMkt);drawHistChart();}},30);
}}

// Tabs de mercado
function showMkt(code){{
  document.querySelectorAll('.mkt-section').forEach(function(s){{s.style.display='none';}});
  var sec=document.getElementById('mkt-'+code);
  if(sec)sec.style.display='';
  document.querySelectorAll('.btn-mkt').forEach(function(b){{b.classList.remove('active');}});
  var btn=document.getElementById('tab-'+code);
  if(btn)btn.classList.add('active');
  _activeMkt=code;
  setTimeout(function(){{drawAllCharts(code);}},30);
}}

// Collapse/expand
function tgC(el){{
  var tr=el.closest('tr');
  var cat=tr.dataset.cat;
  var tbody=tr.closest('tbody');
  var rows=tbody.querySelectorAll('[data-p="'+cat+'"]');
  var op=rows[0]&&rows[0].style.display==='none';
  rows.forEach(function(r){{r.style.display=op?'':'none';}});
  tr.classList.toggle('open',op);
}}
function tgS(el){{
  var tr=el.closest('tr');
  var cat=tr.dataset.p;
  var tbody=tr.closest('tbody');
  var subRows=[];
  var nextEl=tr.nextElementSibling;
  while(nextEl&&!nextEl.classList.contains('rs')&&!nextEl.classList.contains('rc')){{
    if(nextEl.classList.contains('ra')&&nextEl.dataset.p===cat)subRows.push(nextEl);
    nextEl=nextEl.nextElementSibling;
  }}
  if(subRows.length===0){{
    var allRa=tbody.querySelectorAll('tr.ra[data-p="'+cat+'"]');
    var op=true;
    allRa.forEach(function(r){{if(r.style.display!=='none')op=false;}});
    allRa.forEach(function(r){{r.style.display=op?'':'none';}});
    tr.classList.toggle('open',op);
  }}else{{
    var op=subRows[0].style.display==='none';
    subRows.forEach(function(r){{r.style.display=op?'':'none';}});
    tr.classList.toggle('open',op);
  }}
}}
function expAll(code){{document.getElementById('tb-'+code).querySelectorAll('.rs,.ra').forEach(function(r){{r.style.display='';}});}}
function colAll(code){{document.getElementById('tb-'+code).querySelectorAll('.rs,.ra').forEach(function(r){{r.style.display='none';}});}}

// Toggles por mercado — usa registro propio para evitar sobreescritura por hoisting
function togRev(code){{var f=window._mktTogRev&&window._mktTogRev[code];if(f)f();}}
function togU(code){{var f=window._mktTogU&&window._mktTogU[code];if(f)f();}}

// Graficos de mercado
function getCSS(v){{return getComputedStyle(document.documentElement).getPropertyValue(v).trim()||v;}}

function drawWeekChart(canvas, data){{
  if(!canvas||!canvas.offsetWidth)return;
  var dpr=window.devicePixelRatio||1;
  var W=canvas.offsetWidth,H=canvas.offsetHeight||110;
  canvas.width=W*dpr;canvas.height=H*dpr;
  var ctx=canvas.getContext('2d');ctx.scale(dpr,dpr);
  var vals=data.weekTotalsU,n=vals.length;
  var maxV=Math.max.apply(null,vals)*1.15||1;
  var pad={{l:36,r:10,t:8,b:28}};
  var bw=Math.floor((W-pad.l-pad.r)/n*0.55);
  var gap=(W-pad.l-pad.r-bw*n)/(n+1);
  var t3=getCSS('--t3'),bd=getCSS('--bd'),acc=getCSS('--acc'),t2=getCSS('--t2');
  ctx.strokeStyle=bd;ctx.lineWidth=.5;
  for(var i=0;i<=4;i++){{
    var y=pad.t+(H-pad.t-pad.b)/4*i;
    ctx.beginPath();ctx.moveTo(pad.l,y);ctx.lineTo(W-pad.r,y);ctx.stroke();
    if(i<4){{ctx.fillStyle=t3;ctx.font='9px sans-serif';ctx.textAlign='right';ctx.fillText(Math.round(maxV/4*(4-i)),pad.l-4,y+3);}}
  }}
  var bestI=vals.indexOf(Math.max.apply(null,vals));
  vals.forEach(function(v,i){{
    var x=pad.l+gap*(i+1)+bw*i;
    var bh=(v/maxV)*(H-pad.t-pad.b);
    var y=H-pad.b-bh;
    ctx.fillStyle=(i===bestI)?acc:'rgba(255,153,0,.45)';
    ctx.beginPath();if(ctx.roundRect)ctx.roundRect(x,y,bw,bh,3);else ctx.rect(x,y,bw,bh);ctx.fill();
    ctx.fillStyle=(i===bestI)?acc:t2;ctx.font='bold 9px sans-serif';ctx.textAlign='center';
    ctx.fillText(v>=1000?(v/1000).toFixed(1)+'k':v,x+bw/2,y-3);
    ctx.fillStyle=t3;ctx.font='9px sans-serif';
    ctx.fillText(data.weekNames[i].split('–')[0].trim(),x+bw/2,H-pad.b+11);
  }});
}}

function drawCatChart(canvas, data){{
  if(!canvas||!canvas.offsetWidth)return;
  var dpr=window.devicePixelRatio||1;
  var W=canvas.offsetWidth||300,H=canvas.offsetHeight||110;
  canvas.width=W*dpr;canvas.height=H*dpr;
  var ctx=canvas.getContext('2d');ctx.scale(dpr,dpr);
  var isDark=(document.documentElement.getAttribute('data-theme')||'dark')==='dark';
  var cats=data.cats.filter(function(c){{return c.u.some(function(v){{return v>0;}});}});
  var n=cats.length,NWEEKS=4;
  if(n===0)return;
  var NAME_W=76,PAD_T=16,PAD_B=4,PAD_R=6;
  var CELL_W=Math.floor((W-NAME_W-PAD_R)/NWEEKS);
  var CELL_H=Math.floor((H-PAD_T-PAD_B)/n);
  var t2=isDark?'#8892b0':'#556080';
  var t3=isDark?'#4a5270':'#8898b8';
  var bd=isDark?'rgba(255,255,255,.06)':'rgba(0,0,0,.06)';
  ['W1','W2','W3','W4'].forEach(function(lbl,wi){{
    var cx=NAME_W+wi*CELL_W+CELL_W/2;
    ctx.fillStyle=t3;ctx.font='bold 9px sans-serif';ctx.textAlign='center';ctx.fillText(lbl,cx,PAD_T-4);
  }});
  cats.forEach(function(cat,ci){{
    var maxU=Math.max.apply(null,cat.u)||1;
    var ry=PAD_T+ci*CELL_H;
    ctx.fillStyle=t2;ctx.font='10px sans-serif';ctx.textAlign='right';
    var lbl=cat.name.length>10?cat.name.slice(0,9)+'…':cat.name;
    ctx.fillText(lbl,NAME_W-4,ry+CELL_H/2+3.5);
    for(var wi=0;wi<NWEEKS;wi++){{
      var ratio=cat.u[wi]/maxU;
      var alpha=0.12+ratio*0.78;
      ctx.fillStyle='rgba(255,153,0,'+alpha.toFixed(2)+')';
      ctx.fillRect(NAME_W+wi*CELL_W+1,ry+1,CELL_W-2,CELL_H-2);
    }}
    if(ci<n-1){{ctx.strokeStyle=bd;ctx.lineWidth=.5;ctx.beginPath();ctx.moveTo(0,ry+CELL_H);ctx.lineTo(W,ry+CELL_H);ctx.stroke();}}
  }});
}}

function drawAllCharts(code){{
  var data=window.CHART_DATA&&window.CHART_DATA[code];
  if(!data)return;
  var cw=document.getElementById('cvWeek-'+code);
  var cc=document.getElementById('cvCat-'+code);
  if(cw)drawWeekChart(cw,data);
  if(cc)drawCatChart(cc,data);
}}

// Grafico historico
var _histColors=['#ff9900','#22c55e','#60a5fa','#f472b6','#a78bfa','#34d399','#fb923c','#f59e0b','#e879f9'];
function buildHistBtns(){{
  var div=document.getElementById('hist-mkt-btns');
  if(!div)return;
  MKT_LIST.forEach(function(m,i){{
    var b=document.createElement('button');
    b.className='btn-hist'+(m.code==='{default_mkt}'?' on':'');
    b.id='hbtn-'+m.code;
    b.textContent=m.flag+' '+m.code;
    b.onclick=(function(code){{return function(){{
      var idx=_histMkts.indexOf(code);
      if(idx>=0){{_histMkts.splice(idx,1);document.getElementById('hbtn-'+code).classList.remove('on');}}
      else{{_histMkts.push(code);document.getElementById('hbtn-'+code).classList.add('on');}}
      drawHistChart();
    }}}})(m.code);
    div.appendChild(b);
  }});
}}

function drawHistChart(){{
  var canvas=document.getElementById('cvHist');
  if(!canvas||!HIST_DATA||HIST_DATA.length===0)return;
  var dpr=window.devicePixelRatio||1;
  var W=canvas.offsetWidth||800,H=canvas.offsetHeight||130;
  canvas.width=W*dpr;canvas.height=H*dpr;
  var ctx=canvas.getContext('2d');ctx.scale(dpr,dpr);
  var isDark=(document.documentElement.getAttribute('data-theme')||'dark')==='dark';
  var t1=isDark?'#dde3f5':'#131728';
  var t2=isDark?'#8892b0':'#556080';
  var t3=isDark?'#4a5270':'#8898b8';
  var bd=isDark?'#2a2f4a':'#cdd2e8';
  var pad={{l:40,r:16,t:10,b:28}};
  var n=HIST_DATA.length;
  if(n===0)return;
  var allVals=[];
  HIST_DATA.forEach(function(d){{
    _histMkts.forEach(function(c){{allVals.push(d.markets[c]||0);}});
  }});
  var maxV=Math.max.apply(null,allVals)*1.15||10;
  var cw=(W-pad.l-pad.r)/(n>1?n-1:1);
  // grid
  ctx.strokeStyle=bd;ctx.lineWidth=.5;
  for(var gi=0;gi<=4;gi++){{
    var gy=pad.t+(H-pad.t-pad.b)/4*gi;
    ctx.beginPath();ctx.moveTo(pad.l,gy);ctx.lineTo(W-pad.r,gy);ctx.stroke();
    ctx.fillStyle=t3;ctx.font='9px sans-serif';ctx.textAlign='right';
    ctx.fillText(Math.round(maxV/4*(4-gi)),pad.l-4,gy+3);
  }}
  // lines per market
  _histMkts.forEach(function(code,mi){{
    var color=_histColors[MKT_LIST.findIndex(function(m){{return m.code===code;}})]||'#888';
    ctx.strokeStyle=color;ctx.lineWidth=2;ctx.lineJoin='round';
    ctx.beginPath();
    var started=false;
    HIST_DATA.forEach(function(d,i){{
      var v=d.markets[code]||0;
      var x=pad.l+i*cw;
      var y=H-pad.b-(v/maxV)*(H-pad.t-pad.b);
      if(!started){{ctx.moveTo(x,y);started=true;}}else ctx.lineTo(x,y);
    }});
    ctx.stroke();
    // dots
    HIST_DATA.forEach(function(d,i){{
      var v=d.markets[code]||0;
      var x=pad.l+i*cw;
      var y=H-pad.b-(v/maxV)*(H-pad.t-pad.b);
      ctx.fillStyle=color;ctx.beginPath();ctx.arc(x,y,3,0,Math.PI*2);ctx.fill();
    }});
  }});
  // x labels
  HIST_DATA.forEach(function(d,i){{
    if(n<=12||i%Math.ceil(n/10)===0){{
      var x=pad.l+i*cw;
      ctx.fillStyle=t3;ctx.font='8px sans-serif';ctx.textAlign='center';
      ctx.fillText('W'+d.iso,x,H-pad.b+10);
    }}
  }});
  // legend
  var lx=pad.l;
  _histMkts.forEach(function(code,mi){{
    var color=_histColors[MKT_LIST.findIndex(function(m){{return m.code===code;}})]||'#888';
    var mkt=MKT_LIST.find(function(m){{return m.code===code;}});
    ctx.fillStyle=color;ctx.fillRect(lx,pad.t-1,16,3);
    ctx.fillStyle=t2;ctx.font='9px sans-serif';ctx.textAlign='left';
    ctx.fillText((mkt?mkt.flag+' ':'')+code,lx+20,pad.t+6);
    lx+=70;
  }});
}}

// Init
buildHistBtns();
window.addEventListener('load',function(){{
  showMkt('{default_mkt}');
  setTimeout(drawHistChart,60);
}});
window.addEventListener('resize',function(){{drawAllCharts(_activeMkt);drawHistChart();}});
</script>
</body>
</html>'''

with open(OUT_PATH, 'w', encoding='utf-8') as f:
    f.write(html)
print(f'\n✓ Dashboard: {OUT_PATH} ({len(html):,} bytes)')
print(f'=== RESUMEN ===')
for mc in MARKET_CONFIGS:
    md = all_market_data.get(mc['code'])
    if md is None:
        print(f'  {mc["flag"]} {mc["code"]}: sin datos')
    else:
        print(f'  {mc["flag"]} {mc["code"]}: {md["n_active"]} ASINs · {round(md["total_u"]):,} uds · {round(md["total_r"]/1000,1)}k{mc["currency"]}')
